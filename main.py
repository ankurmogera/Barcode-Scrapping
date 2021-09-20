from openpyxl import load_workbook
from bs4 import BeautifulSoup
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
import time


def website1():
    workbook = load_workbook(filename="Barcodes_not_found.xlsx")
    sheet = workbook.active
    count = 1
    for col in sheet['A']:
        url = 'https://www.upcitemdb.com/upc/{}'.format(col.value)
        options = Options()
        options.add_argument('--headless')
        options.add_argument('--disable-gpu')  # Last I checked this was necessary.
        driver = webdriver.Chrome('./chromedriver', options=options)
        driver.get(url)
        time.sleep(3)
        html = driver.page_source
        soup = BeautifulSoup(html, 'html.parser')
        # div = soup.find('ol', {'class': 'num'})
        try:
            for li in soup.find('ol', {'class': 'num'}):
                sheet.cell(row=count, column=3).value = "{}".format(li.text)
                workbook.save("Barcodes_not_found.xlsx")
                count += 1
                print(li.text)
                print(count)
                break
            driver.close()
        except:
            workbook.save("Barcodes_not_found.xlsx")
            count += 1
            print(count)
            driver.close()

def website2():
    workbook = load_workbook(filename="Barcodes_not_found.xlsx")
    sheet = workbook.active
    count = 1
    for col in sheet['A']:
        url = 'https://www.barcodelookup.com/{}'.format(col.value)
        driver = webdriver.Chrome('./chromedriver')
        driver.minimize_window()
        driver.get(url)
        time.sleep(3)
        html = driver.page_source
        soup = BeautifulSoup(html, 'html.parser')
        # div = soup.find('ol', {'class': 'num'})
        try:
            div = soup.find('img', {'id': 'img_preview'})
            print(div['alt'])
            sheet.cell(row=count, column=4).value = "{}".format(div['alt'])
            workbook.save("Barcodes_not_found.xlsx")
            count += 1
            driver.close()
        except:
            workbook.save("Barcodes_not_found.xlsx")
            count += 1
            print(count)
            driver.close()


def main():
    print("Choose a website for barcode lookup:")
    print("1. www.upcitemdb.com")
    print("2. www.barcodelookup.com")
    option = input("Enter number: ")
    if option == str(1):
        website1()
    elif option == str(2):
        website2()
    else:
        exit()

main()


